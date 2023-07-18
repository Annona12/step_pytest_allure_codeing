# 开发者：Annona
# 开发时间：2023/7/11 9:21
import logging
import time
import allure
import openpyxl
import pytest

from constant.constant import *
from data_driver.deal_excel import read_excel
from data_driver.deal_xml import set_xml, set_xml_string
from tools.tools import Tools


def setup_module(module):
    global tools, all_val, excel_path_list
    tools = Tools()
    excel_path_list = EXCEL_PATH_LIST
    all_val = {}
    MY_LOGGER.info('setup_module前置函数执行完成，开始执行测试用例！！！')


@pytest.mark.parametrize('data',read_excel())
def test_jys(data):
    # 记录日志
    MY_LOGGER.info('从excel获取的测试用例数据！！！！')
    MY_LOGGER.info(data)
    if data[12] == 'no':
        pytest.skip("标记该用例为不执行")
    if data[-1] is not None:
        allure.dynamic.feature(data[-1])
    if data[1] is not None:
        allure.dynamic.story(data[1])
    if data[2] is not None:
        allure.dynamic.title(data[2])
    if data[3] is not None:
        allure.dynamic.description(data[3])
    if data[4] is not None:
        allure.dynamic.severity(data[4])

    excel = openpyxl.load_workbook(data[-2])
    sheet = excel[data[-1]]
    row_num = data[0] + 1

    action_str = str(data[5])
    xml_str = data[6]
    param_str = data[7]

    act_list = action_str.split(';')
    xml_list = xml_str.split(';')
    param_list = param_str.split(';')

    # 判断用例中传递的功能号、读取报文地址、修改参数个数是否对应正确
    if len(act_list) == len(xml_list) == len(param_list):
        for i in range(len(act_list)):
            long_date, local_date, local_time = tools.get_system_time()
            try:
                param_list_i = eval(param_list[i])
            except Exception as e:
                MY_LOGGER.error('用例中传递的功能号、读取报文地址、修改参数个数不正确请检查！！！')
                MY_LOGGER.error(str(e))
                sheet.cell(row_num, 12).value = '修改参数转换成字典格式错误，请检查修改参数列数据！！！'
                excel.save(data[-2])
                assert 0==1
            else:
                param_list_i['sendDateTime'] = long_date
                param_list_i['ordDate'] = local_date
                param_list_i['ordTime'] = local_time
                param_list_i['actDate'] = local_date
                param_list_i['actTime'] = local_time
                param_list_i['orgOrdDate'] = local_date
                # 按照时间戳生成主键sysOrdID，并将该值存储在all_val，以便后续查询数字库，做断言使用
                if 'sysOrdID' in param_list_i.keys():
                    sysord_id = str(int(time.time()))
                    param_list_i['sysOrdID'] = sysord_id
                    all_val['sysOrdID'] = sysord_id

                # 如果参数中有orgSysOrdID字段，则需要取出all_val中的sysOrdID值，对该笔报价进行撤单
                if 'orgSysOrdID' in param_list_i.keys():
                    param_list_i['orgSysOrdID'] = all_val['sysOrdID']
                # 获取了所有最新的需要修改的参数之后，调用工具函数将各个参数写入xml文件中
                set_xml(xml_list[i], param_list_i)
                data_xml = set_xml_string('data/temp.xml')
                MY_LOGGER.info(data_xml)
                # 发送请求
                result = tools.send_post(act_list[i], data_xml)
                MY_LOGGER.info(result)
                time.sleep(5)

        if data[8] is not None and data[9] is not None and data[10] is not None:
            sql_str = data[8]
            sql_str_list = sql_str.split(';')

            hope_result = data[10]
            hope_result_list = hope_result.split(';')
            hope_result_dic_list = []

            sql_param_str = data[9]
            sql_param_list = sql_param_str.split(';')
            param_list_num = []
            length = len(sql_param_list)

            for j in range(length):
                try:
                    param_list_num.append(eval(sql_param_list[j]))
                except:
                    sheet.cell(row_num, 12).value = '数据库变量转换失败，请检查数据库变量列的数据！！！'
                    MY_LOGGER.info('数据库变量转换失败，请检查数据库变量列的数据！！！')
                    excel.save(data[-2])
                    assert 0 == 1
                else:
                    try:
                        for item in hope_result_list:
                            temp_item = eval(item)
                            hope_result_dic_list.append(temp_item)
                    except:
                        sheet.cell(row_num, 12).value = '预期结果值转换成字典格式失败，请检查预期结果列数据！！！'
                        MY_LOGGER.info('预期结果值转换成字典格式失败，请检查预期结果列数据！！！')
                        excel.save(data[-2])
                        assert 0 == 1
                    else:
                        if len(hope_result_dic_list) == len(sql_str_list):
                            for i in range(len(sql_str_list)):
                                tools.conn_oracle(
                                    dsn=ORACLE_DSN,
                                    user=ORACLE_USER,
                                    passwd=ORACLE_PASSWORD
                                )
                                try:
                                    sql = sql_str_list[i].format(*param_list_num)
                                    MY_LOGGER.info(sql)
                                except:
                                    sheet.cell(row_num, 12).value = 'sql语句有误，请检查！！！'
                                    MY_LOGGER.info('sql语句有误，请检查！！！')
                                    excel.save(data[-2])
                                    assert 0 == 1
                                else:
                                    result_list = tools.sql_check(sql)
                                    MY_LOGGER.info(result_list)
                                    for key in hope_result_dic_list[i].keys():
                                        real_result = result_list[key]
                                        if real_result == hope_result_dic_list[i][key]:
                                            sheet.cell(row_num, 12).value = '测试通过'
                                            MY_LOGGER.info('测试通过')
                                            excel.save(data[-2])
                                        else:
                                            sheet.cell(row_num, 12).value = '测试不通过'
                                            MY_LOGGER.info('测试不通过')
                                            excel.save(data[-2])
                                    for key in hope_result_dic_list[i].keys():
                                        with allure.step(f"数据库{key}断言"):
                                            with allure.step(f'预期结果：{key}={hope_result_dic_list[i][key]}'):
                                                real_result = result_list[key]
                                            with allure.step(f'实际结果：{key}={real_result}'):
                                                assert real_result == hope_result_dic_list[i][key]
                        else:
                            sheet.cell(row_num, 12).value = '查询sql与预期结果字典个数不对应，请检查！！！'
                            MY_LOGGER.info('查询sql与预期结果字典个数不对应，请检查！！！')
                            excel.save(data[-2])
                            assert 0 == 1
        else:
            sheet.cell(row_num, 12).value = "查询数据库sql或者预期结果确实，请检查！！！"
            MY_LOGGER.info('查询数据库sql或者预期结果或者数据变量缺失，请检查！！！')
            excel.save(data[-2])
            assert 0 == 1
    else:
        sheet.cell(row_num, 12).value = "用例中传递的功能号、读取报文地址、修改参数个数不对应请检查！！！"
        MY_LOGGER.error('用例中传递的功能号、读取报文地址、修改参数个数不正确请检查！！！')
        excel.save(data[-2])
        assert 0==1
