# 开发者：Annona
# 开发时间：2023/7/11 9:42
# 开发者：Annona
# 开发时间：2023/6/16 17:36
import logging

import openpyxl

from constant.constant import EXCEL_PATH_LIST
from tools.tools import Tools


def read_excel():
    global tools
    tools = Tools()
    excel_path_list = EXCEL_PATH_LIST
    all_case_list = []
    try:
        for excel_path in excel_path_list:
            excel = openpyxl.load_workbook(excel_path)
            for sheet_name in excel.sheetnames:
                sheet = excel[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    if isinstance(row[0],int):
                        all_case_list.append(row+(excel_path,sheet_name))
        tools.my_logger().info('读取所有excel用例完成！！！')
    except Exception as e:
        tools.my_logger().info(str(e))
    return all_case_list
# testcase = read_excel()
# print(testcase)
# print(testcase[0])
# print(len(testcase))

